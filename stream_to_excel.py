"""
Simulador de fluxo de datos de sensores -> Excel
Formato do dato: [temperatura, humidade, altura, presión, timestamp]

Follas do Excel:
  · Rexistro     — todas as lecturas en bruto
  · Temperatura  — táboa de seguimento con estatísticas acumuladas
  · Humidade     — ídem
  · Altura       — ídem
  · Presión      — ídem
  · Resumo       — últimas estatísticas globais de cada parámetro
"""

import random
import time
import threading
import queue
import os
from datetime import datetime
from dataclasses import dataclass

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

EXCEL_FILE = "sensores.xlsx"
DATA_QUEUE = queue.Queue()
RUNNING    = True


# ---------------------------------------------------------------------------
# Modelo de dato
# ---------------------------------------------------------------------------

@dataclass
class Lectura:
    temperatura: float   # °C      rango: -10 .. 50
    humidade:    float   # %       rango:   0 .. 100
    altura:      float   # m       rango:   0 .. 3000
    presion:     float   # hPa     rango: 870 .. 1085
    timestamp:   str


# Definición de cada parámetro: (nome folla, unidade, cor cabeceira, atributo en Lectura)
PARAMETROS = [
    ("Temperatura", "°C",  "C0504D", "temperatura"),
    ("Humidade",    "%",   "4472C4", "humidade"),
    ("Altura",      "m",   "9BBB59", "altura"),
    ("Presión",     "hPa", "7030A0", "presion"),
]


# ---------------------------------------------------------------------------
# Produtor: simula o fluxo de datos do sensor
# ---------------------------------------------------------------------------

def xerar_lectura() -> Lectura:
    """Xera unha lectura aleatoria dentro dos rangos típicos de cada sensor."""
    return Lectura(
        temperatura = round(random.uniform(-10.0,  50.0),   1),
        humidade    = round(random.uniform(  0.0, 100.0),   1),
        altura      = round(random.uniform(  0.0, 3000.0),  0),
        presion     = round(random.uniform(870.0, 1085.0),  1),
        timestamp   = datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    )


def produtor(intervalo: int = 5):
    """Emite unha nova lectura cada `intervalo` segundos e métea na cola."""
    global RUNNING
    while RUNNING:
        lectura = xerar_lectura()
        print(
            f"[SENSOR]  {lectura.timestamp}"
            f"  T={lectura.temperatura:>5}°C"
            f"  H={lectura.humidade:>5}%"
            f"  Alt={lectura.altura:>6}m"
            f"  P={lectura.presion:>7}hPa"
        )
        DATA_QUEUE.put(lectura)
        time.sleep(intervalo)


# ---------------------------------------------------------------------------
# Axudantes de estilo
# ---------------------------------------------------------------------------

ALIGN_CTR = Alignment(horizontal="center")
FONT_BOLD = Font(bold=True)


def _recheo(cor: str) -> PatternFill:
    return PatternFill("solid", fgColor=cor)


def _tipo_cabeceira(cor: str) -> Font:
    return Font(bold=True, color=cor)


def _aplicar_fila_cabeceira(ws, cabeceiras: list, cor: str, fila: int = 1):
    """Aplica estilo de cabeceira á fila indicada da folla."""
    recheo = _recheo(cor)
    tipo   = _tipo_cabeceira("FFFFFF")
    for col, texto in enumerate(cabeceiras, 1):
        c = ws.cell(row=fila, column=col, value=texto)
        c.font      = tipo
        c.fill      = recheo
        c.alignment = ALIGN_CTR


# ---------------------------------------------------------------------------
# Gráfica de liñas por parámetro
# ---------------------------------------------------------------------------

def _actualizar_grafica(ws, nome: str, unidade: str, cor: str):
    """Recrea a gráfica de liñas na folla do parámetro."""
    max_fila = ws.max_row
    if max_fila < 3:          # necesítanse polo menos 2 lecturas para trazar liña
        return

    ws._charts = []           # eliminar a gráfica anterior antes de recrear

    chart = LineChart()
    chart.title          = f"{nome} ({unidade})"
    chart.style          = 2
    chart.height         = 14    # cm
    chart.width          = 26    # cm
    chart.y_axis.title   = f"{nome} ({unidade})"
    chart.x_axis.title   = "Lectura #"
    chart.y_axis.numFmt  = "0.0"
    # Evitar que se amontoen as etiquetas do eixo X con moitos puntos
    chart.x_axis.tickLblSkip = max(1, (max_fila - 2) // 12)

    # Serie 1 — valor real
    ref_val = Reference(ws, min_col=3, min_row=1, max_row=max_fila)
    chart.add_data(ref_val, titles_from_data=True)

    # Serie 2 — media acumulada
    ref_media = Reference(ws, min_col=6, min_row=1, max_row=max_fila)
    chart.add_data(ref_media, titles_from_data=True)

    # Categorías (número de lectura, col A)
    cats = Reference(ws, min_col=1, min_row=2, max_row=max_fila)
    chart.set_categories(cats)

    # Estilo serie 0: valor real (cor do parámetro, liña sólida 2 pt)
    s0 = chart.series[0]
    s0.graphicalProperties.line.solidFill = cor
    s0.graphicalProperties.line.width     = 25400   # 2 pt en EMU
    s0.smooth = True

    # Estilo serie 1: media (gris, liña descontinua 1.25 pt)
    s1 = chart.series[1]
    s1.graphicalProperties.line.solidFill = "808080"
    s1.graphicalProperties.line.width     = 15875   # 1.25 pt en EMU
    s1.graphicalProperties.line.prstDash  = "dash"
    s1.smooth = True

    ws.add_chart(chart, "H2")


# ---------------------------------------------------------------------------
# Inicialización do Excel
# ---------------------------------------------------------------------------

def _crear_folla_rexistro(wb):
    """Crea a folla principal con todas as lecturas en bruto."""
    ws = wb.active
    ws.title = "Rexistro"
    cabeceiras = ["#", "Timestamp", "Temperatura (°C)", "Humidade (%)", "Altura (m)", "Presión (hPa)"]
    _aplicar_fila_cabeceira(ws, cabeceiras, "2E75B6")
    for i, ancho in enumerate([6, 20, 18, 13, 12, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho


def _crear_folla_parametro(wb, nome: str, unidade: str, cor: str):
    """Crea a folla de seguimento dun parámetro con estatísticas acumuladas."""
    ws = wb.create_sheet(nome)
    cabeceiras = [
        "#",
        "Timestamp",
        f"{nome} ({unidade})",
        "Mín acum.",
        "Máx acum.",
        "Media acum.",
    ]
    _aplicar_fila_cabeceira(ws, cabeceiras, cor)
    for i, ancho in enumerate([6, 20, 18, 12, 12, 13], 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho


def _crear_folla_resumo(wb):
    """Crea a folla de resumo con estatísticas globais de todos os parámetros."""
    ws = wb.create_sheet("Resumo")
    ws["A1"] = "Resumo de parámetros"
    ws["A1"].font = Font(bold=True, size=13)

    cabeceiras = ["Parámetro", "Último valor", "Mínimo", "Máximo", "Media", "Total lecturas", "Actualizado"]
    _aplicar_fila_cabeceira(ws, cabeceiras, "2E75B6", fila=3)

    for fila, (nome, unidade, _, _) in enumerate(PARAMETROS, 4):
        ws.cell(row=fila, column=1, value=f"{nome} ({unidade})").font = FONT_BOLD

    for i, ancho in enumerate([20, 14, 10, 10, 10, 15, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho


def inicializar_excel():
    """Crea o ficheiro Excel coa estrutura inicial; se xa existe, cárgao."""
    if os.path.exists(EXCEL_FILE):
        print(f"[EXCEL]   Cargando '{EXCEL_FILE}' existente.")
        return

    wb = openpyxl.Workbook()
    _crear_folla_rexistro(wb)
    for nome, unidade, cor, _ in PARAMETROS:
        _crear_folla_parametro(wb, nome, unidade, cor)
    _crear_folla_resumo(wb)

    wb.save(EXCEL_FILE)
    print(f"[EXCEL]   Ficheiro '{EXCEL_FILE}' creado.")


# ---------------------------------------------------------------------------
# Actualización do Excel con cada nova lectura
# ---------------------------------------------------------------------------

def actualizar_excel(lectura: Lectura):
    wb = load_workbook(EXCEL_FILE)

    # --- 1. Folla Rexistro ---
    ws_reg  = wb["Rexistro"]
    num_reg = ws_reg.max_row       # fila 1 = cabeceira, polo que max_row = nº de rexistros
    sig_reg = ws_reg.max_row + 1

    for col, valor in enumerate([
        num_reg,
        lectura.timestamp,
        lectura.temperatura,
        lectura.humidade,
        lectura.altura,
        lectura.presion,
    ], 1):
        ws_reg.cell(row=sig_reg, column=col, value=valor)

    # --- 2. Follas de seguimento por parámetro ---
    for nome, unidade, cor, attr in PARAMETROS:
        ws    = wb[nome]
        valor = getattr(lectura, attr)

        ultima_fila = ws.max_row  # fila 1 = cabeceira

        if ultima_fila == 1:
            # Primeira lectura: as estatísticas acumuladas son o propio valor
            min_acum  = valor
            max_acum  = valor
            media_acum = valor
            n          = 1
        else:
            # Ler estatísticas da última fila para actualización incremental
            prev_min   = ws.cell(row=ultima_fila, column=4).value
            prev_max   = ws.cell(row=ultima_fila, column=5).value
            prev_media = ws.cell(row=ultima_fila, column=6).value
            n_prev     = ultima_fila - 1   # número de lecturas anteriores
            n          = n_prev + 1

            min_acum   = min(prev_min, valor)
            max_acum   = max(prev_max, valor)
            media_acum = round((prev_media * n_prev + valor) / n, 2)

        nova_fila = ultima_fila + 1
        for col, v in enumerate([n, lectura.timestamp, valor, min_acum, max_acum, media_acum], 1):
            ws.cell(row=nova_fila, column=col, value=v)

        _actualizar_grafica(ws, nome, unidade, cor)

    # --- 3. Folla Resumo ---
    ws_res = wb["Resumo"]
    for fila_res, (nome, _, _, attr) in enumerate(PARAMETROS, 4):
        ws_par     = wb[nome]
        ultima     = ws_par.max_row
        ultimo_val = ws_par.cell(row=ultima, column=3).value
        min_v      = ws_par.cell(row=ultima, column=4).value
        max_v      = ws_par.cell(row=ultima, column=5).value
        media_v    = ws_par.cell(row=ultima, column=6).value
        total      = ultima - 1

        for col, v in enumerate([
            None,        # columna A xa ten o label
            ultimo_val,
            min_v,
            max_v,
            media_v,
            total,
            lectura.timestamp,
        ], 1):
            if v is not None:
                ws_res.cell(row=fila_res, column=col, value=v)

    wb.save(EXCEL_FILE)
    print(
        f"[EXCEL]   #{num_reg:>4}"
        f"  T={lectura.temperatura:>5}°C"
        f"  H={lectura.humidade:>5}%"
        f"  Alt={lectura.altura:>6}m"
        f"  P={lectura.presion:>7}hPa"
        f"  -> gardado"
    )


# ---------------------------------------------------------------------------
# Consumidor: le a cola e actualiza o Excel
# ---------------------------------------------------------------------------

def consumidor():
    global RUNNING
    while RUNNING or not DATA_QUEUE.empty():
        try:
            lectura = DATA_QUEUE.get(timeout=1)
            actualizar_excel(lectura)
        except queue.Empty:
            continue
        except Exception as e:
            print(f"[ERRO]    {e}")


# ---------------------------------------------------------------------------
# Principal
# ---------------------------------------------------------------------------

def main():
    global RUNNING
    print("=" * 60)
    print("  Simulador de sensores  ->  Excel")
    print("  Follas: Rexistro | Temperatura | Humidade | Altura | Presión | Resumo")
    print(f"  Ficheiro: {EXCEL_FILE}   |   Intervalo: 5 s")
    print("  Ctrl+C para deter")
    print("=" * 60)

    inicializar_excel()

    t_prod = threading.Thread(target=produtor,   args=(5,), daemon=True)
    t_cons = threading.Thread(target=consumidor, daemon=True)

    t_prod.start()
    t_cons.start()

    try:
        while True:
            time.sleep(0.5)
    except KeyboardInterrupt:
        print("\n[INFO]    Detendo...")
        RUNNING = False
        time.sleep(2)
        print("[INFO]    Programa rematado. Datos en:", EXCEL_FILE)


if __name__ == "__main__":
    main()
